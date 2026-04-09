"""Excel management service – read, write, and chart research data."""

from __future__ import annotations

import io
import logging
from datetime import datetime
from pathlib import Path
from typing import Any

import matplotlib
matplotlib.use("Agg")  # Non-interactive backend (macOS / headless)
import matplotlib.pyplot as plt
import openpyxl
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from config import ExcelConfig

log = logging.getLogger(__name__)

# Colour palette
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
ALT_FILL = PatternFill("solid", fgColor="D6E4F0")

SHEET_COLUMNS: dict[str, list[str]] = {
    "Macro Themes": ["ID", "Theme", "Description", "Source Count", "Updated"],
    "Meso Themes": ["ID", "Theme", "Parent Macro", "Description", "Source Count", "Updated"],
    "Micro Themes": ["ID", "Theme", "Parent Meso", "Description", "Details", "Updated"],
    "Articles": ["ID", "Title", "Authors", "Year", "DOI", "URL", "Abstract", "Macro", "Meso", "Micro", "Added"],
    "Patents": ["ID", "Title", "Patent ID", "Assignee", "Inventor", "Date", "URL", "Abstract", "Macro", "Meso", "Micro", "Added"],
    "Search Results": ["ID", "Query", "Title", "URL", "Snippet", "Source Type", "Added"],
}


# ─── Workbook helpers ─────────────────────────────────────────────────────────

def ensure_workbook() -> openpyxl.Workbook:
    """Open existing workbook or create a new one with all required sheets."""
    path = ExcelConfig.workbook_path()
    path.parent.mkdir(parents=True, exist_ok=True)

    if path.exists():
        wb = openpyxl.load_workbook(str(path))
    else:
        wb = openpyxl.Workbook()
        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    # Ensure all sheets exist
    for sheet_name in SHEET_COLUMNS:
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
            _write_header(ws, SHEET_COLUMNS[sheet_name])

    wb.save(str(path))
    return wb


def _write_header(ws: openpyxl.worksheet.worksheet.Worksheet, columns: list[str]) -> None:
    """Write a styled header row."""
    for col_idx, header in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = max(len(header) + 4, 12)
    ws.row_dimensions[1].height = 20


def _next_id(ws: openpyxl.worksheet.worksheet.Worksheet) -> int:
    """Return next available row ID (max existing + 1)."""
    max_id = 0
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        if isinstance(row[0], int) and row[0] > max_id:
            max_id = row[0]
    return max_id + 1


def _append_row(ws: openpyxl.worksheet.worksheet.Worksheet, values: list[Any]) -> int:
    """Append a row to the worksheet and return its row number."""
    row_num = ws.max_row + 1
    for col_idx, value in enumerate(values, start=1):
        cell = ws.cell(row=row_num, column=col_idx, value=value)
        if row_num % 2 == 0:
            cell.fill = ALT_FILL
    return row_num


# ─── Public API ───────────────────────────────────────────────────────────────

def save_article(article: dict[str, Any], classification: dict[str, Any] | None = None) -> int:
    """Save an article to the Articles sheet. Returns the assigned ID."""
    wb = ensure_workbook()
    ws = wb["Articles"]
    new_id = _next_id(ws)
    themes = classification.get("themes", {}) if classification else {}
    macro = ", ".join(themes.get("macro", []))
    meso = ", ".join(themes.get("meso", []))
    micro = ", ".join(themes.get("micro", []))
    _append_row(
        ws,
        [
            new_id,
            article.get("title", ""),
            article.get("authors", ""),
            article.get("year", ""),
            article.get("doi", ""),
            article.get("url", ""),
            article.get("abstract", "")[:500] if article.get("abstract") else "",
            macro,
            meso,
            micro,
            datetime.now().strftime("%Y-%m-%d %H:%M"),
        ],
    )
    wb.save(str(ExcelConfig.workbook_path()))
    return new_id


def save_patent(patent: dict[str, Any], classification: dict[str, Any] | None = None) -> int:
    """Save a patent to the Patents sheet. Returns the assigned ID."""
    wb = ensure_workbook()
    ws = wb["Patents"]
    new_id = _next_id(ws)
    themes = classification.get("themes", {}) if classification else {}
    macro = ", ".join(themes.get("macro", []))
    meso = ", ".join(themes.get("meso", []))
    micro = ", ".join(themes.get("micro", []))
    _append_row(
        ws,
        [
            new_id,
            patent.get("title", ""),
            patent.get("patent_id", ""),
            patent.get("assignee", ""),
            patent.get("inventor", ""),
            patent.get("date", ""),
            patent.get("url", ""),
            patent.get("abstract", "")[:500] if patent.get("abstract") else "",
            macro,
            meso,
            micro,
            datetime.now().strftime("%Y-%m-%d %H:%M"),
        ],
    )
    wb.save(str(ExcelConfig.workbook_path()))
    return new_id


def save_search_result(result: dict[str, Any], query: str) -> int:
    """Save a search result row. Returns the assigned ID."""
    wb = ensure_workbook()
    ws = wb["Search Results"]
    new_id = _next_id(ws)
    _append_row(
        ws,
        [
            new_id,
            query,
            result.get("title", ""),
            result.get("url", ""),
            result.get("snippet", ""),
            result.get("type", "web"),
            datetime.now().strftime("%Y-%m-%d %H:%M"),
        ],
    )
    wb.save(str(ExcelConfig.workbook_path()))
    return new_id


def save_theme(theme_type: str, theme: str, parent: str = "", description: str = "") -> int:
    """Save a macro/meso/micro theme. Returns the assigned ID."""
    sheet_map = {"macro": "Macro Themes", "meso": "Meso Themes", "micro": "Micro Themes"}
    sheet_name = sheet_map.get(theme_type.lower(), "Macro Themes")
    wb = ensure_workbook()
    ws = wb[sheet_name]
    new_id = _next_id(ws)
    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    if theme_type.lower() == "macro":
        _append_row(ws, [new_id, theme, description, 0, now])
    elif theme_type.lower() == "meso":
        _append_row(ws, [new_id, theme, parent, description, 0, now])
    else:  # micro
        _append_row(ws, [new_id, theme, parent, description, "", now])
    wb.save(str(ExcelConfig.workbook_path()))
    return new_id


def get_sheet_data(sheet_name: str) -> list[dict[str, Any]]:
    """Read all rows from a sheet as list of dicts."""
    path = ExcelConfig.workbook_path()
    if not path.exists():
        ensure_workbook()

    wb = openpyxl.load_workbook(str(path), read_only=True)
    if sheet_name not in wb.sheetnames:
        return []

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(h) if h is not None else "" for h in rows[0]]
    result = []
    for row in rows[1:]:
        if any(v is not None for v in row):
            result.append(dict(zip(headers, row)))
    return result


def get_all_sheets_summary() -> dict[str, int]:
    """Return row count for each sheet."""
    path = ExcelConfig.workbook_path()
    if not path.exists():
        ensure_workbook()
    wb = openpyxl.load_workbook(str(path), read_only=True)
    return {name: max(0, ws.max_row - 1) for name, ws in ((n, wb[n]) for n in wb.sheetnames)}


def generate_chart(chart_type: str = "bar", topic: str = "") -> io.BytesIO:
    """
    Generate a matplotlib chart of the current research data distribution.
    Returns a BytesIO PNG buffer.

    chart_type: 'bar', 'pie', or 'line'
    """
    summary = get_all_sheets_summary()
    labels = list(summary.keys())
    values = list(summary.values())

    fig, ax = plt.subplots(figsize=(10, 6))
    title = f"Research Data Overview{' – ' + topic if topic else ''}"

    if chart_type == "pie":
        non_zero = [(l, v) for l, v in zip(labels, values) if v > 0]
        if non_zero:
            l2, v2 = zip(*non_zero)
            ax.pie(v2, labels=l2, autopct="%1.1f%%", startangle=140)
        ax.set_title(title)
    elif chart_type == "line":
        ax.plot(labels, values, marker="o", linewidth=2, color="#1F4E79")
        ax.fill_between(range(len(labels)), values, alpha=0.2, color="#1F4E79")
        ax.set_xticks(range(len(labels)))
        ax.set_xticklabels(labels, rotation=30, ha="right")
        ax.set_ylabel("Count")
        ax.set_title(title)
    else:  # bar (default)
        bars = ax.bar(labels, values, color=["#1F4E79", "#ED7D31", "#A9D18E", "#FF0000", "#7030A0", "#00B0F0"])
        ax.set_xticks(range(len(labels)))
        ax.set_xticklabels(labels, rotation=30, ha="right")
        ax.set_ylabel("Count")
        ax.set_title(title)
        for bar, val in zip(bars, values):
            ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 0.1, str(val), ha="center", va="bottom")

    plt.tight_layout()
    buf = io.BytesIO()
    plt.savefig(buf, format="png", dpi=120)
    plt.close(fig)
    buf.seek(0)
    return buf


def generate_themes_chart(topic: str = "") -> io.BytesIO:
    """Generate a stacked bar chart showing themes distribution."""
    macro_data = get_sheet_data("Macro Themes")
    meso_data = get_sheet_data("Meso Themes")
    micro_data = get_sheet_data("Micro Themes")

    labels = ["Macro", "Meso", "Micro"]
    counts = [len(macro_data), len(meso_data), len(micro_data)]

    fig, ax = plt.subplots(figsize=(8, 5))
    colors = ["#1F4E79", "#ED7D31", "#A9D18E"]
    bars = ax.bar(labels, counts, color=colors, width=0.5)

    for bar, count in zip(bars, counts):
        ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 0.1, str(count), ha="center")

    title = f"Theme Distribution{' – ' + topic if topic else ''}"
    ax.set_title(title)
    ax.set_ylabel("Count")
    plt.tight_layout()

    buf = io.BytesIO()
    plt.savefig(buf, format="png", dpi=120)
    plt.close(fig)
    buf.seek(0)
    return buf


def workbook_path() -> Path:
    return ExcelConfig.workbook_path()
